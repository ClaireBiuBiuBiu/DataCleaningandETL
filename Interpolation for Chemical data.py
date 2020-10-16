import os
import csv
import pandas as pd
import time
import xlrd
import numpy as np
import openpyxl as op
import os.path
from scipy import interpolate
os.chdir("my working director")
book  = op.load_workbook('filename.xlsx')
sheets = book.sheetsnames

headings = []
headings.append('filename')
headings.append('material')

for i in range(7,25):
    headings.append(str(book[sheets[1]].cell(i,1).value).strip())
headings.append('A')
headings.append('B')
...

#Interpolation data point
xx = [10,50,100,200,400,600,1000]

def check_boudary():
    for n_limit in range(4,sh.max_row+1):
        if not sh.cell(n_limit,5).value:
            break
    return n_limit
test = []
test.append(headings)

for i in range(1,len(sheets)):
    sh = book[sheets[i]]
    n_limt = check_boudary()
    cell = sh.cell(1,1).value
    Filename = str(cell)
    cell2 = sh.cell(4,2).value
    Liquid = str(cell2)

    rows = []
    rows.append(Filename)
    rows.append(Liquid)
    for i in range(7,25):
        rows.append(str(sh.cell(i,2).value).strip())

    x = []
    for i in range(4,n_limt):
        x.append(sh.cell(i,5).value)
    y = []
    for i in range(4,n_limt):
        y.append(sh.cell(i,7).value)
    yy = interpolate.interpid(x,y,kind = 'linear',fill_value = 'extrapolate')

    zz = []
    zz.append(sh.cell(4,10).value)
    zz.append(sh.cell(n_limt-1,10).value)

    zz.append(sh.cell(4,8).value)
    zz.apend(sh.cell(n_limt-1,8).value)

    zz.append(sh.cell(4,9).value)
    zz.append(sh.cell(n_limt-1,9).value)

    data_matrix = rows + yy + zz
    test.append(data_matrix)
with open('output.csv','w',newline = "") as f:
    writer = csv.writer(f)
    writer.writerows(test)
    f.close()

#Below is the visualization for linear interpolation

import matplotlib.pyplot as plt
plt.scatter(x,y)
plt.plot(xx,yy)
plt.show()